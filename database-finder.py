import streamlit as st
import pandas as pd
import streamlit.components.v1 as components
from openpyxl import load_workbook


# Cache the workbook loading to avoid unnecessary reloads
@st.cache_data
def load_workbook_data():
    wb = load_workbook("data/database_checklist.xlsx")
    return wb.active


# Load the Excel file
ws = load_workbook_data()

# Extract headers and data from openpyxl
header_row = [cell.value for cell in ws[1]]
content_types = [val for val in header_row[1:] if val is not None]
data_start_row = 3

# Streamlit UI
st.title("Business Database Finder")

# Label and dropdown on the same line to avoid extra padding
st.subheader("Select content types:", divider=False, width="stretch")
selected_types = st.multiselect(
    "Select content types", content_types, label_visibility="collapsed"
)

if selected_types:
    # Match columns for selected content types
    matching_cols = [i for i, val in enumerate(header_row) if val in selected_types]

    # Extract names and hyperlinks from Excel directly
    names = []
    urls = []

    for row in ws.iter_rows(min_row=3, min_col=1, max_col=1):
        cell = row[0]
        names.append(cell.value)
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        else:
            urls.append("")

    data_rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))
    df = pd.DataFrame(data_rows)
    content_data = df.iloc[:, matching_cols].copy()
    content_data.columns = selected_types
    result = pd.DataFrame({"Database": names, "URL": urls})
    result = pd.concat([result, content_data.reset_index(drop=True)], axis=1)

    # Filter for rows that match all selected types
    for col in selected_types:
        if col == "OTHER":
            result = result[
                result[col].notna() & result[col].astype(str).str.strip().ne("")
            ]
        else:
            result = result[
                result[col].notna()
                & result[col].astype(str).str.strip().str.lower().eq("y")
            ]
    result = result.reset_index(drop=True)
    result = result.apply(
        lambda col: col.map(
            lambda x: "✅" if isinstance(x, str) and x.strip().lower() == "y" else x
        )
    )

    result["Database"] = [
        (
            f"[{name}]({url})"
            if isinstance(url, str) and ("http" in url or "www." in url)
            else name
        )
        for name, url in zip(result["Database"], result["URL"])
    ]
    result.drop(columns=["URL"], inplace=True)

    # Display output with proper formatting using markdown to render links
    st.write(
        f"### Databases matching **all** selected content types ({len(result)} found):"
    )
    st.markdown(result.to_markdown(index=False), unsafe_allow_html=True)
else:
    st.info("⬆️ Select your content types above.")
